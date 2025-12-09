from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from common.config.common.settings import settings
from logging import Logger
from common.infrastructure.logging.config import get_logger

_LOGGER: Logger = get_logger(__name__)


class SurveyPdfGenerator:
    """
    Utility responsible for generating PDF files for surveys based on the Excel template
    shipped with the project. The generator copies the template's content and appends a
    flattened summary of the survey detail payload so every PDF reflects the stored data.
    """

    def __init__(
        self,
        template_path: Optional[Path | str] = None,
        output_dir: Optional[Path | str] = None,
    ):
        self._project_root = Path(__file__).resolve().parents[4]
        self._template_path = Path(template_path) if template_path else self._project_root / "ADR.Formularios.Visita.xlsx"
        self._output_dir = Path(output_dir) if output_dir else self._project_root / "images" / "pdf"
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._template_cache: Optional[List[Tuple[str, List[List[str]]]]] = None

    def file_exists(self, relative_path: Optional[str]) -> bool:
        if not relative_path:
            return False
        candidate = (self._project_root / relative_path).resolve()
        return candidate.exists()

    def generate_pdf(
        self,
        survey_type: int,
        survey_id: int,
        survey_payload: Dict[str, Any],
    ) -> str:
        filename = f"survey_{survey_type}_{survey_id}.pdf"
        output_path = self._output_dir / filename
        self._create_pdf(output_path, survey_payload or {})
        relative_path = output_path.relative_to(self._project_root)
        return relative_path.as_posix()

    @staticmethod
    def build_public_url(relative_path: str) -> str:
        cleaned_path = relative_path.lstrip("/")
        base_url = settings.backend_url.rstrip("/")
        return f"{base_url}/{cleaned_path}"

    def _create_pdf(self, output_path: Path, payload: Dict[str, Any]) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover - dependency issue caught at runtime
            _LOGGER.error("openpyxl is required to generate survey PDFs: %s", exc)
            raise

        try:
            from reportlab.lib import colors  # type: ignore[import-not-found]
            from reportlab.lib.pagesizes import A4  # type: ignore[import-not-found]
            from reportlab.lib.styles import getSampleStyleSheet  # type: ignore[import-not-found]
            from reportlab.platypus import (  # type: ignore[import-not-found]
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:  # pragma: no cover
            _LOGGER.error("reportlab is required to generate survey PDFs: %s", exc)
            raise

        styles = getSampleStyleSheet()
        document = SimpleDocTemplate(str(output_path), pagesize=A4, title="Formulario de Visita")
        elements: List[Any] = [
            Paragraph("Formulario de Visita", styles["Title"]),
            Spacer(1, 12),
        ]

        flattened_payload = self._flatten_payload(payload)
        if flattened_payload:
            elements.append(Paragraph("Resumen de la encuesta", styles["Heading2"]))
            for key, value in flattened_payload:
                elements.append(Paragraph(f"<b>{key}:</b> {value}", styles["Normal"]))
            elements.append(Spacer(1, 12))

        template_sections = self._load_template_sections(load_workbook)
        for sheet_name, rows in template_sections:
            if not rows:
                continue
            elements.append(Paragraph(f"Plantilla: {sheet_name}", styles["Heading2"]))
            table = Table(self._normalize_rows(rows))
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 16))

        if not template_sections:
            elements.append(
                Paragraph(
                    "No fue posible cargar la plantilla ADR.Formularios.Visita.xlsx.",
                    styles["Normal"],
                )
            )

        document.build(elements)

    def _load_template_sections(
        self,
        load_workbook: Any,
    ) -> List[Tuple[str, List[List[str]]]]:
        if self._template_cache is not None:
            return self._template_cache

        if not self._template_path.exists():
            _LOGGER.warning("Template Excel file not found at %s", self._template_path)
            self._template_cache = []
            return self._template_cache

        workbook = load_workbook(self._template_path, data_only=True)
        sections: List[Tuple[str, List[List[str]]]] = []
        for worksheet in workbook.worksheets:
            rows: List[List[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                normalized_row = [self._stringify_cell(cell) for cell in row]
                if any(cell for cell in normalized_row):
                    rows.append(normalized_row)
            sections.append((worksheet.title, rows))
        self._template_cache = sections
        return sections

    @staticmethod
    def _stringify_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (str, int, float)):
            return str(value)
        return json.dumps(value, ensure_ascii=False)

    @staticmethod
    def _normalize_rows(rows: Sequence[Sequence[str]]) -> List[List[str]]:
        if not rows:
            return []
        max_len = max(len(row) for row in rows)
        normalized: List[List[str]] = []
        for row in rows:
            padded = list(row) + [""] * (max_len - len(row))
            normalized.append(padded)
        return normalized

    def _flatten_payload(
        self,
        payload: Any,
        prefix: str = "",
        max_items: int = 80,
    ) -> List[Tuple[str, str]]:
        flattened: List[Tuple[str, str]] = []

        def _walk(value: Any, current_prefix: str) -> None:
            if len(flattened) >= max_items:
                return
            if isinstance(value, dict):
                for key, nested_value in value.items():
                    new_prefix = f"{current_prefix}.{key}" if current_prefix else str(key)
                    _walk(nested_value, new_prefix)
            elif isinstance(value, list):
                for index, nested_value in enumerate(value):
                    new_prefix = f"{current_prefix}[{index}]"
                    _walk(nested_value, new_prefix)
            else:
                key = current_prefix or "valor"
                flattened.append((key, self._stringify_cell(value)))

        _walk(payload, prefix)
        return flattened
